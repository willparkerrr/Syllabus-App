import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_excel(assignments, file_path):
    """Creates a formatted Excel file from AI-extracted assignments, removing description column."""
    if not assignments:
        raise ValueError("ERROR: No assignments found to export.")

    try:
        # Convert to DataFrame
        df = pd.DataFrame(assignments)

        # Ensure required columns exist
        expected_columns = ["course", "assignment", "due_date"]
        for col in expected_columns:
            if col not in df.columns:
                df[col] = "Unknown"

        # Convert due_date to datetime format for sorting
        df["due_date"] = pd.to_datetime(df["due_date"], errors="coerce")

        # Sort assignments by due_date and course
        df = df.sort_values(by=["due_date", "course"], ascending=[True, True])

        # Remove the "description" column if it exists
        if "description" in df.columns:
            df.drop(columns=["description"], inplace=True)

        # Save to Excel
        df.to_excel(file_path, index=False, sheet_name="Master Schedule")

        # Load the workbook to apply formatting
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Apply header formatting (bold text, centered)
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")
        for col in range(1, len(df.columns) + 1):
            sheet.cell(row=1, column=col).font = header_font
            sheet.cell(row=1, column=col).alignment = header_alignment

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2  # Adjust padding

        # Apply alternating row colors for better readability
        fill_color1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_color2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row in range(2, sheet.max_row + 1):
            fill = fill_color1 if row % 2 == 0 else fill_color2
            for col in range(1, len(df.columns) + 1):
                sheet.cell(row=row, column=col).fill = fill

        # Save formatted Excel file
        workbook.save(file_path)
        print(f"Formatted Master Excel file created: {file_path}")

    except Exception as e:
        print(f"ERROR: Failed to generate Excel file: {e}")
        raise
